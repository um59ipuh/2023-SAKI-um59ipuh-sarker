# crawler for extracting data from https://www.kba.de/DE/Statistik/Fahrzeuge/Neuzulassungen/Umwelt/n_umwelt_node.html
# 
# craw all past data till the end of given year

import requests
import datetime
import json
from openpyxl import Workbook, load_workbook
import csv

# get all infos for extracting data
def get_info():
    # load json for info
    json_file = "info.json"
    with open(json_file) as f:
        infos = json.load(f)
    # extract mobi infos
    return infos["sources"][1]

def extract_sales_data(filename):
    # load xlsx file
    exfile = load_workbook(f"{filename}.xlsx")
    # get specific sheet
    sales_sheet = exfile["FZ 28.9"]
    
    temp_data = [] # {"state": [], "sales": [], "month":[], "year": []}
    
    cell_value = sales_sheet.cell(row=13, column=2).value.split(" ")
    months_de = ["januar", "februar", "märz", "april", "mai", "juni", "juli", "august", "september", 
                 "oktober", "november", "dezember"]
    
    month, year = cell_value[0], cell_value[1]
    
    # make data dictionary
    if month.lower() in months_de:
        for row in sales_sheet.iter_rows(min_row=14, max_row=30, min_col=2, max_col=3, values_only=True):
            single_row = {}
            single_row["state"] = row[0]
            single_row["sales"] = row[1]
            single_row["month"] = month
            single_row["year"] = year
            temp_data.append(single_row)
            
    # write data to csv
    csv_file = f"{filename}.csv"
    with open(csv_file, 'w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=temp_data[0].keys())
        # Write the header
        writer.writeheader()
        for row in range(1, len(temp_data)):
            # Write the data
            writer.writerow(temp_data[row])
        print(f"{csv_file} - created successfully.")

def generate_filename(date = ""):
    if date != "":
        return f"kba_sales_data_{date}"
    # Save the response content to a file
    current_time = datetime.datetime.now()
    # Format the timestamp as a string
    timestamp = current_time.strftime("%Y_%m_%d")
    file_name = f"kba_sales_data_{timestamp}"
    return file_name

def crawl_past_data(date, version):
    # get properties from json
    infos = get_info()
    url = infos["absolute_url"].format(date, version)

    # get the file from online
    response = requests.get(url)
    
    # Check if the request was successful (status code 200)
    if response.status_code == 200:
        # save the file with this filename
        file_woext = f"files/{generate_filename(date)}"
        file_wext = f"{file_woext}.xlsx"
        # open and save contents
        with open(file_wext, 'wb') as file:
            file.write(response.content)
        print(f"{file_wext} - File downloaded successfully.")
        
        # now extract exact data and convert to csv file
        extract_sales_data(file_woext)
    else:
        print(f"{file_wext} - Failed to download the file.")


def prev_year_month(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1

def crawl_sales_data_till(year, month):
    fixed_dates = {"2023 4":("28_2023_04", "2"), "2023 3":("28_2023_03", "6"), "2023 2":("28_2023_02", "6"), "2023 1": ("28_2023_01", "6"), "2022 12": ("28_2022_12", "6"), "2022 11": ("28_2022_11", "7"), 
                   "2022 10": ("28_2022_10", "4"), "2022 9": ("28_2022_09", "4"), "2022 8": ("28_2022_08", "5"), "2022 7":("28_2022_07", "6"), "2022 6": ("28_2022_06", "8")}
    
    curr_ = (2023, 4)
    while True:
        date, version = fixed_dates[f"{curr_[0]} {curr_[1]}"]
        crawl_past_data(date, version)
        if curr_[0] == year and curr_[1] == month:
            break
        else:
            curr_ = prev_year_month(curr_[0], curr_[1])

# crawl new data if avaliable
def crawl_new_data():
    data = get_info()
    #TODO: write code for extract file if new available
    pass

def crawl():
    
    # TODO::
    # crawl all past data if not available
    crawl_sales_data_till(2022, 6)
    # crawl new data if available
        
if __name__ == '__main__':
    crawl()